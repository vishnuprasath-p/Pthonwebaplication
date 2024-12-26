from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import sqlite3
from openpyxl import Workbook
from fpdf import FPDF
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Database connection function
def connect_db():
    return sqlite3.connect('user_data.db')

# Alter table to add 'settled' column
def alter_table():
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_data_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            title TEXT,
            amount REAL NOT NULL,
            total_time REAL NOT NULL,
            settled INTEGER NOT NULL DEFAULT 0  -- Add settled column
        )
    ''')
    cursor.execute('''
        INSERT INTO user_data_new (id, date, title, amount, total_time, settled)
        SELECT id, date, title, amount, total_time, 0 FROM user_data
    ''')
    connection.commit()
    cursor.execute('DROP TABLE user_data')
    cursor.execute('ALTER TABLE user_data_new RENAME TO user_data')
    connection.commit()
    connection.close()

alter_table()

# Create table if it doesn't exist
def create_table():
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            title TEXT NOT NULL,
            amount REAL NOT NULL,
            total_time REAL NOT NULL,
            settled INTEGER NOT NULL DEFAULT 0  -- Add settled column
        )
    ''')
    connection.commit()
    connection.close()

create_table()

# Date validation function
def validate_date(date_text):
    try:
        datetime.strptime(date_text, "%Y-%m-%d")
        return True
    except ValueError:
        return False

# Home route
@app.route('/')
def index():
    connection = connect_db()
    cursor = connection.cursor()
    
    # Fetch unsettled data
    cursor.execute('SELECT * FROM user_data WHERE settled = 0')
    rows = cursor.fetchall()

    # Fetch settled data
    cursor.execute('SELECT * FROM user_data WHERE settled = 1')
    settled_rows = cursor.fetchall()

    # Calculate total settled amount
    cursor.execute('SELECT SUM(amount) FROM user_data WHERE settled = 1')
    total_settled_amount = cursor.fetchone()[0] or 0

    connection.close()
    
    return render_template('index.html', rows=rows, settled_rows=settled_rows, total_settled_amount=total_settled_amount)

# Add new data route
@app.route('/add', methods=['POST'])
def add():
    date = request.form['date']
    title = request.form['title']
    amount = request.form['amount']
    total_time = request.form['total_time']

    if not date or not title or not amount or not total_time:
        flash('All fields are required!')
        return redirect(url_for('index'))

    if not validate_date(date):
        flash('Invalid date format! Use YYYY-MM-DD.')
        return redirect(url_for('index'))

    try:
        amount = float(amount)
        total_time = float(total_time)
    except ValueError:
        flash('Amount and Total Time must be numeric!')
        return redirect(url_for('index'))

    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('''
        INSERT INTO user_data (date, title, amount, total_time, settled)
        VALUES (?, ?, ?, ?, ?)
    ''', (date, title, amount, total_time, 0))  # Default 'settled' is 0
    connection.commit()
    connection.close()

    flash('Data saved successfully!')
    return redirect(url_for('index'))

# Delete data route
@app.route('/delete/<int:id>')
def delete(id):
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('DELETE FROM user_data WHERE id = ?', (id,))
    connection.commit()
    connection.close()
    flash('Data deleted successfully!')
    return redirect(url_for('index'))

# Settle data route
@app.route('/settle', methods=['POST'])
def settle():
    settle_ids = request.form.getlist('settle_ids')
    
    if not settle_ids:
        flash('No items selected to settle!')
        return redirect(url_for('index'))

    connection = connect_db()
    cursor = connection.cursor()
    
    # Update the settled status for the selected items
    cursor.execute('''
        UPDATE user_data
        SET settled = 1
        WHERE id IN ({})
    '''.format(','.join('?' * len(settle_ids))), tuple(settle_ids))
    connection.commit()
    connection.close()

    flash('Selected items have been marked as settled!')
    return redirect(url_for('index'))

# Filter data by date range
@app.route('/filter', methods=['POST'])
def filter_data():
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    
    if not validate_date(start_date) or not validate_date(end_date):
        flash('Invalid date format! Use YYYY-MM-DD.')
        return redirect(url_for('index'))

    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM user_data WHERE date BETWEEN ? AND ?', (start_date, end_date))
    rows = cursor.fetchall()
    connection.close()

    return render_template('index.html', rows=rows)

# Month-wise total route
@app.route('/month_summary')
def month_summary():
    connection = connect_db()
    cursor = connection.cursor()
    
    # Calculate the month-wise total amount
    cursor.execute('''
        SELECT strftime('%Y-%m', date) AS month, SUM(amount) AS total_amount
        FROM user_data
        GROUP BY month
        ORDER BY month
    ''')
    monthly_totals = cursor.fetchall()
    connection.close()

    return render_template('index.html', monthly_totals=monthly_totals)

# Export to Excel
def generate_file_name(prefix, extension):
    today = datetime.today().strftime('%Y-%m-%d')
    counter = 1
    while True:
        file_name = f"{prefix}_{today}_{counter}.{extension}"
        if not os.path.exists(file_name):
            break
        counter += 1
    return file_name

@app.route('/export/excel')
def export_to_excel():
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM user_data')
    rows = cursor.fetchall()
    connection.close()

    if not rows:
        flash('No data to export!')
        return redirect(url_for('index'))

    file_name = generate_file_name("user_data", "xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "Date", "Title", "Amount", "Total Time", "Settled"])
    for row in rows:
        sheet.append(row)

    workbook.save(file_name)
    return send_file(file_name, as_attachment=True)

# Export to PDF
@app.route('/export/pdf')
def export_to_pdf():
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM user_data')
    rows = cursor.fetchall()
    connection.close()

    if not rows:
        flash('No data to export!')
        return redirect(url_for('index'))

    file_name = generate_file_name("user_data", "pdf")
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Add a page to the PDF
    pdf.add_page()
    
    # Set font for the title/header
    pdf.set_font("Arial", 'B', size=16)
    pdf.cell(200, 10, txt="User Data Report", ln=True, align='C')  # Main heading
    pdf.ln(10)  # Line break
    
    # Set font for the column headers
    pdf.set_font("Arial", 'B', size=12)
    pdf.cell(30, 10, "ID", border=1, align='C')
    pdf.cell(40, 10, "Date", border=1, align='C')
    pdf.cell(50, 10, "Title", border=1, align='C')
    pdf.cell(40, 10, "Amount", border=1, align='C')
    pdf.cell(50, 10, "Total Time", border=1, align='C')
    pdf.cell(40, 10, "Settled", border=1, align='C')
    pdf.ln(10)  # Line break

    # Set font for the content
    pdf.set_font("Arial", size=10)

    # Loop through the data and add rows to the PDF
    for row in rows:
        pdf.cell(30, 10, str(row[0]), border=1, align='C')
        pdf.cell(40, 10, str(row[1]), border=1, align='C')
        pdf.cell(50, 10, str(row[2]), border=1, align='C')
        pdf.cell(40, 10, f"{row[3]:.2f}", border=1, align='C')
        pdf.cell(50, 10, f"{row[4]:.2f}", border=1, align='C')
        pdf.cell(40, 10, "Yes" if row[5] == 1 else "No", border=1, align='C')
        pdf.ln(10)  # Line break for each row

    # Save the PDF to a file
    pdf.output(file_name)
    return send_file(file_name, as_attachment=True)

# Retrieve all data route
@app.route('/retrieve_all')
def retrieve_all():
    connection = connect_db()
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM user_data')
    rows = cursor.fetchall()
    connection.close()
    
    return render_template('index.html', rows=rows)

if __name__ == '__main__':
    app.run(debug=True)
